# Create your views here.
import os

from rest_framework.generics import ListAPIView, CreateAPIView, ListCreateAPIView
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from .models import HostCategory, Host
from .serializers import HostCategoryModelSeiralizer, HostModelSerializers
from rest_framework.permissions import IsAuthenticated
from openpyxl import load_workbook
from io import BytesIO


class HostCategoryListAPIView(ListCreateAPIView):
    """主机类别"""
    queryset = HostCategory.objects.filter(is_show=True, is_deleted=False).order_by("orders", "-id").all()
    serializer_class = HostCategoryModelSeiralizer
    permission_classes = [IsAuthenticated]


class HostModelViewSet(ModelViewSet):
    serializer_class = HostModelSerializers
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # 重写qureyset方法，补充过滤主机列表参数，获取主机列表
        category_id = self.request.query_params.get("category", None)
        queryset = Host.objects
        if category_id is not None:
            queryset = queryset.filter(category_id=category_id)

        return queryset.all()


def read_host_excel_data(io_data, default_password=''):
    """
    从excel中读取主机列表信息
    io_data: 主机列表的字节流
    default_password: 主机的默认登录密码
    """

    # 加载某一个excel文件
    wb = load_workbook(io_data)
    # 获取worksheet对象的两种方式
    worksheet = wb.worksheets[1]

    # c1 = worksheet.cell(2, 1)  # 第二行第一列
    # print("c1 data:::", c1.value)

    # 查询出数据库现有的所有分类数据[ID，name]
    # 由于拿到的是分类名称，所以我们要找到对应名称的分类id，才能去数据库里面存储
    category_list = HostCategory.objects.values_list('id', 'name')

    # 主机列表
    host_info_list = []
    for row in worksheet.iter_rows(2):
        if not row[0].value: continue
        one_row_dict = {}  # 单个主机信息字典

        for category_data in category_list:
            # print(category_data[1],type(category_data[1]),category,type(category))
            if category_data[1].strip() == row[0].value:
                one_row_dict['category'] = category_data[0]
                break

        one_row_dict["name"] = row[1].value  # 主机别名
        one_row_dict['ip_addr'] = row[2].value  # 主机地址
        one_row_dict['port'] = row[3].value  # 主机端口号
        one_row_dict['username'] = row[4].value  # 登录账户名

        excel_pwd = row[5].value
        try:
            pwd = str(excel_pwd)  # 这样强转容易报错，最好捕获一下异常，并记录单元格位置，给用户保存信息时，可以提示用户哪个单元格的数据有问题
        except Exception as e:
            pwd = default_password

        if not pwd.strip():
            pwd = default_password

        one_row_dict['password'] = pwd
        one_row_dict['description'] = row[6].value
        print("one_row_dict", one_row_dict)

        host_info_list.append(one_row_dict)

    # 校验主机数据
    # 将做好的主机信息字典数据通过我们添加主机时的序列化器进行校验
    res_data = {}  # 存放上传成功之后需要返回的主机数据和某些错误信息数据
    serializers_host_res_data = []
    res_error_data = []
    for k, host_data in enumerate(host_info_list):
        # 反序列化校验每一个主机信息
        serializer = HostModelSerializers(data=host_data)
        if serializer.is_valid():
            new_host_obj = serializer.save()
            serializers_host_res_data.append(new_host_obj)
        else:
            # 报错，并且错误信息中应该体验错误的数据位置
            res_error_data.append({'error': f'该表{k + 1}行数据有误,其他没有问题的数据，已经添加成功了，请求失败数据改完之后，重新上传这个错误数据，成功的数据不需要上传！'})

    # # 再次调用序列化器进行数据的序列化，返回给客户端
    serializer = HostModelSerializers(instance=serializers_host_res_data, many=True)
    res_data['data'] = serializer.data
    res_data['error'] = res_error_data

    return res_data


class ExcelHostView(APIView):
    def post(self, request):
        """批量导入主机列表"""
        # 接受客户端上传的数据
        host_excel = request.FILES.get("host_excel")
        default_password = request.data.get("default_password")
        # print("host_excel:::", host_excel, type(host_excel))
        # print("default_password:::", default_password, type(default_password))
        path = os.path.join("upload_files", host_excel.name)
        # # 把上传文件全部写入到字节流，就不需要保存到服务端硬盘了。
        io_data = BytesIO()
        for line in host_excel:
            io_data.write(line)

        #
        data = read_host_excel_data(io_data, default_password)
        #
        return Response(data)
